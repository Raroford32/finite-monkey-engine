import pandas as pd
from tqdm import tqdm
import json
from openai_api.openai import analyze_code_assumptions, ask_deepseek, extract_structured_json
import concurrent.futures
from threading import Lock
import math
import re
import time
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

class ResProcessor:
    def __init__(self, df, max_group_size=10, iteration_rounds=2, enable_chinese_translation=False):
        """
        Initialize ResProcessor for advanced exploit discovery analysis
        
        Args:
            df: DataFrame containing vulnerability data
            max_group_size: Maximum vulnerabilities per group, default 10
            iteration_rounds: Number of iteration rounds, default 2
            enable_chinese_translation: Enable Chinese translation, default False (force English)
        """
        self.df = df
        self.lock = Lock()
        self.max_group_size = max_group_size
        self.iteration_rounds = iteration_rounds
        # Force English-only output for exploit discovery
        self.enable_chinese_translation = False  # Always disable Chinese translation
        
        print(f"ResProcessor initialized for exploit discovery:")
        print(f"  - Max group size: {self.max_group_size}")
        print(f"  - Iteration rounds: {self.iteration_rounds}")
        print(f"  - Output language: English only (Chinese translation disabled)")

    def process(self):
        """Main processing function for multi-round vulnerability aggregation"""
        print("Starting vulnerability aggregation processing for exploit discovery...")
        print(f"Total vulnerability count: {len(self.df)}")
        
        # Add auxiliary columns
        self.df['flow_code_len'] = self.df['业务流程代码'].str.len()
        
        # First step: Group by business flow code
        initial_groups = list(self.df.groupby('业务流程代码'))
        print(f"Initial group count: {len(initial_groups)}")
        
        # Print initial grouping details
        print("\n=== Initial Grouping Details ===")
        for i, (flow_code, group) in enumerate(initial_groups):
            flow_code_preview = flow_code[:100] + "..." if len(flow_code) > 100 else flow_code
            print(f"分组 {i+1}: 业务流程代码长度={len(flow_code)}, 漏洞数量={len(group)}")
            print(f"  代码预览: {flow_code_preview}")
            if len(group) > self.max_group_size:
                print(f"  ⚠️  该分组超过最大限制({self.max_group_size})，需要细分")
        
        # 第二步：细分大组（超过10个漏洞的组）
        refined_groups = self._refine_large_groups(initial_groups)
        print(f"\n细分后分组数量: {len(refined_groups)}")
        
        # 第三步：多轮迭代归集
        current_groups = refined_groups
        for round_num in range(self.iteration_rounds):
            print(f"\n{'='*50}")
            print(f"开始第 {round_num + 1} 轮归集")
            print(f"{'='*50}")
            print(f"输入分组数量: {len(current_groups)}")
            
            # 打印当前轮次分组详情
            total_vulns = sum(len(group) for group in current_groups)
            print(f"当前轮次总漏洞数量: {total_vulns}")
            for i, group in enumerate(current_groups):
                print(f"  分组 {i+1}: {len(group)} 个漏洞")
            
            current_groups = self._iteration_round(current_groups, round_num + 1)
            
            # 检查返回的数据类型并处理提前停止
            if isinstance(current_groups, list) and len(current_groups) > 0:
                if isinstance(current_groups[0], dict):
                    # 返回的是字典列表，说明是最终结果
                    print(f"第 {round_num + 1} 轮归集完成，返回最终结果: {len(current_groups)} 个漏洞报告")
                    break
                else:
                    # 返回的是DataFrame列表，继续下一轮
                    print(f"第 {round_num + 1} 轮归集后分组数量: {len(current_groups)}")
            else:
                print(f"第 {round_num + 1} 轮归集完成，无有效结果")
                break
        
        # 第四步：构建最终结果
        final_results = self._build_final_results(current_groups)
        print(f"\n最终结果: {len(final_results)} 个漏洞报告")
        
        # 第五步：中文翻译（可选）
        final_results = self._translate_to_chinese(final_results)
        
        # 清理辅助列并返回结果
        new_df = pd.DataFrame(final_results)
        if 'flow_code_len' in new_df.columns:
            new_df = new_df.drop('flow_code_len', axis=1)
            
        original_columns = [col for col in self.df.columns if col != 'flow_code_len']
        new_df = new_df[original_columns]
        
        return new_df

    def _refine_large_groups(self, initial_groups):
        """细分大组，确保每组不超过最大限制"""
        refined_groups = []
        
        print("\n=== 开始细分大组 ===")
        
        # 使用多线程处理大组细分
        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
            future_to_group = {
                executor.submit(self._process_single_group, i, flow_code, group): (i, flow_code, group)
                for i, (flow_code, group) in enumerate(initial_groups)
            }
            
            with tqdm(total=len(initial_groups), desc="细分大组") as pbar:
                for future in concurrent.futures.as_completed(future_to_group):
                    try:
                        subgroups = future.result()
                        refined_groups.extend(subgroups)
                    except Exception as e:
                        i, flow_code, group = future_to_group[future]
                        print(f"分组 {i+1} 细分失败: {str(e)}")
                        refined_groups.append(group)
                    pbar.update(1)
        
        print(f"细分完成: {len(initial_groups)} 个初始分组 -> {len(refined_groups)} 个细分后分组")
        return refined_groups

    def _process_single_group(self, index, flow_code, group):
        """处理单个分组的细分"""
        if len(group) <= self.max_group_size:
            flow_code_preview = flow_code[:50] + "..." if len(flow_code) > 50 else flow_code
            print(f"分组 {index+1} (业务流程代码: {flow_code_preview}): 大小 {len(group)} <= {self.max_group_size}，无需细分")
            return [group]
        else:
            # 将大组拆分为小组
            num_subgroups = math.ceil(len(group) / self.max_group_size)
            flow_code_preview = flow_code[:50] + "..." if len(flow_code) > 50 else flow_code
            print(f"分组 {index+1} (业务流程代码: {flow_code_preview}): 大小 {len(group)} > {self.max_group_size}，需要拆分为 {num_subgroups} 个子组")
            
            group_list = group.to_dict('records')
            subgroups = []
            
            for j in range(num_subgroups):
                start_idx = j * self.max_group_size
                end_idx = min(start_idx + self.max_group_size, len(group_list))
                subgroup_data = group_list[start_idx:end_idx]
                subgroup_df = pd.DataFrame(subgroup_data)
                subgroups.append(subgroup_df)
                print(f"  子组 {j+1}: 漏洞数量 {len(subgroup_data)} (行 {start_idx+1}-{end_idx})")
            
            return subgroups

    def _iteration_round(self, groups, round_num):
        """执行一轮迭代归集"""
        print(f"\n--- 第 {round_num} 轮处理开始 ---")
        print(f"处理 {len(groups)} 个分组")
        
        # 打印每个分组的详情
        for i, group in enumerate(groups):
            print(f"  输入分组 {i+1}: {len(group)} 个漏洞")
        
        # 步骤1：对每个组进行漏洞分类
        print(f"\n步骤1: 开始漏洞分类...")
        classified_groups = []
        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
            future_to_group = {
                executor.submit(self._classify_vulnerabilities_in_group, group): (group, i) 
                for i, group in enumerate(groups)
            }
            
            with tqdm(total=len(groups), desc=f"第{round_num}轮-漏洞分类") as pbar:
                for future in concurrent.futures.as_completed(future_to_group):
                    try:
                        classified_result = future.result()
                        original_group, group_index = future_to_group[future]
                        
                        print(f"  分组 {group_index+1} 分类结果: {len(original_group)} -> {len(classified_result)} 个子组")
                        for j, subgroup in enumerate(classified_result):
                            print(f"    子组 {j+1}: {len(subgroup)} 个漏洞")
                        
                        classified_groups.extend(classified_result)
                    except Exception as e:
                        original_group, group_index = future_to_group[future]
                        print(f"  分组 {group_index+1} 分类失败，保持原组: {str(e)}")
                        classified_groups.append(original_group)
                    pbar.update(1)
        
        print(f"分类完成: {len(groups)} 个输入分组 -> {len(classified_groups)} 个分类后分组")
        
        # 步骤2：对分类后的组进行去重处理（每组只保留第一个漏洞）
        print(f"\n步骤2: 开始去重处理...")
        deduplicated_results = []
        
        for i, group in enumerate(classified_groups):
            # 输入验证：确保group是DataFrame
            if not isinstance(group, pd.DataFrame):
                print(f"  分组 {i+1}: ⚠️  输入不是DataFrame，类型: {type(group)}")
                if isinstance(group, dict):
                    # 如果是单个字典，直接保留
                    print(f"  分组 {i+1}: 单个字典，直接保留")
                    deduplicated_results.append(group)
                    continue
                elif isinstance(group, list):
                    # 如果是字典列表，转换为DataFrame
                    group = pd.DataFrame(group)
                else:
                    print(f"  分组 {i+1}: ❌ 无法处理的输入类型，跳过")
                    continue
            
            if len(group) <= 1:
                # 单个漏洞，直接保留
                print(f"  分组 {i+1}: 单个漏洞，直接保留")
                preserved_vuln = group.iloc[0].to_dict()
                deduplicated_results.append(preserved_vuln)
            else:
                # 多个漏洞，只保留第一个
                first_vuln = group.iloc[0]
                removed_count = len(group) - 1
                print(f"  分组 {i+1}: {len(group)} 个漏洞 -> 保留第1个，删除 {removed_count} 个重复项")
                
                # 记录被删除的漏洞ID
                removed_ids = [str(row['ID']) for _, row in group.iloc[1:].iterrows()]
                print(f"    删除的漏洞ID: {', '.join(removed_ids)}")
                print(f"    保留的漏洞ID: {first_vuln['ID']}")
                
                preserved_vuln = first_vuln.to_dict()
                deduplicated_results.append(preserved_vuln)
        
        print(f"去重完成: {len(classified_groups)} 个分类后分组 -> {len(deduplicated_results)} 个去重后结果")
        
        # 统计去重效果
        original_total = sum(len(group) for group in classified_groups)
        final_total = len(deduplicated_results)
        removed_total = original_total - final_total
        print(f"去重统计: 原始 {original_total} 个漏洞 -> 最终 {final_total} 个漏洞，删除了 {removed_total} 个重复项")
        
        # 检查是否需要提前停止
        if final_total < 40:
            print(f"\n🎯 归集结果已少于25个({final_total}个)，提前停止迭代")
            print(f"第 {round_num} 轮为提前结束轮次")
            return deduplicated_results  # 返回字典列表作为最终结果
        
        # 步骤3：准备下一轮的分组（每10个一组）
        if round_num < self.iteration_rounds:
            print(f"\n步骤3: 准备第 {round_num+1} 轮分组...")
            next_round_groups = self._prepare_next_round_groups_simple(deduplicated_results)
            print(f"为下一轮准备了 {len(next_round_groups)} 个分组")
            
            # 打印下一轮分组详情
            total_items = sum(len(group) for group in next_round_groups)
            print(f"下一轮总项目数: {total_items}")
            for i, group in enumerate(next_round_groups):
                print(f"  下一轮分组 {i+1}: {len(group)} 个项目")
            
            return next_round_groups  # 返回DataFrame列表继续下一轮
        else:
            print(f"\n第 {round_num} 轮为最后一轮，不需要准备下一轮分组")
            return deduplicated_results  # 返回字典列表作为最终结果

    def _classify_vulnerabilities_in_group(self, group):
        """使用新的prompt对单个组进行漏洞分类"""
        # 输入验证：确保group是DataFrame
        if not isinstance(group, pd.DataFrame):
            print(f"    ⚠️  输入不是DataFrame，类型: {type(group)}")
            if isinstance(group, dict):
                # 如果是单个字典，转换为DataFrame
                group = pd.DataFrame([group])
            elif isinstance(group, list):
                # 如果是字典列表，转换为DataFrame
                group = pd.DataFrame(group)
            else:
                print(f"    ❌ 无法处理的输入类型，直接返回")
                return [group]
        
        if len(group) <= 1:
            print(f"    单个漏洞无需分类，直接返回")
            return [group]
        
        print(f"    开始分类 {len(group)} 个漏洞...")
        
        # 构建漏洞信息
        vuln_keys = []
        vuln_descriptions = []
        
        for _, row in group.iterrows():
            key = str(row['UUID'])
            description = f"漏洞内容：{row['漏洞结果']}"
            vuln_keys.append(key)
            vuln_descriptions.append(f"Key: {key}\n{description}")
        
        print(f"    构建了 {len(vuln_keys)} 个漏洞的描述信息")
        
        classification_prompt = f"""# Role: 网络安全漏洞分类专家

## Profile
- language: 中文
- description: 专业的网络安全漏洞识别与分类专家，具备深厚的安全技术背景和丰富的漏洞分析经验，能够准确识别各类安全漏洞的本质特征，并进行精准分类
- background: 拥有多年网络安全从业经验，精通各种漏洞类型、攻击原理和防护机制，参与过大量漏洞挖掘、分析和修复工作
- personality: 严谨细致、逻辑清晰、专业专注、追求精确

## Skills

1. 漏洞识别与分析
   - 漏洞本质判断: 能够透过表面现象识别漏洞的根本原理和成因
   - 漏洞类型归纳: 精通OWASP、CVE、CWE等标准分类体系
   - 攻击向量分析: 深度理解各种攻击手段和利用方式
   - 影响范围评估: 准确判断漏洞的危害程度和影响范围

2. 分类与标准化
   - 标准化分类: 基于国际标准和行业最佳实践进行分类
   - 相似性识别: 快速识别不同表述下的相同本质漏洞
   - 层次化归类: 建立清晰的分类层次和逻辑关系
   - 结果格式化: 按照指定格式准确输出分类结果

## Rules

1. 分类基本原则:
   - 本质相同原则: 只有漏洞的根本原理和成因完全相同才归为一类
   - 严格区分原则: 表面相似但本质不同的漏洞必须分为不同类别
   - 完整性原则: 确保所有提供的漏洞key都被正确分类，不遗漏不重复
   - 一致性原则: 采用统一的分类标准，确保分类结果的一致性

2. 分析行为准则:
   - 深度分析: 深入分析每个漏洞的技术细节和攻击原理
   - 对比验证: 通过多维度对比确认漏洞的相似性和差异性
   - 客观判断: 基于技术事实进行分类，避免主观臆断
   - 精准识别: 准确识别漏洞的核心特征和关键差异点

3. 输出限制条件:
   - 格式严格: 严格按照规定的JSON格式输出分类结果
   - 内容纯净: 仅输出分类结果，不包含任何解释说明或额外信息
   - 标识唯一: 确保每个分类组有唯一的分组标识
   - 无冗余信息: 不输出分析过程、理由说明或其他辅助内容

## Workflows

- 目标: 将输入的漏洞按照本质特征进行精准分类
- 步骤 1: 接收并解析所有漏洞key，理解每个漏洞的技术特征和攻击原理
- 步骤 2: 基于漏洞的根本成因、攻击机制、利用方式等核心要素进行深度分析
- 步骤 3: 识别具有相同本质特征的漏洞，将其归为同一类别，为每个类别分配唯一标识
- 步骤 4: 按照标准JSON格式输出最终分类结果
- 预期结果: 输出标准化的漏洞分类结果，格式为{{"group_1":["key1","key2"],"group_2":["key3","key4"],"group_3":["key5"]}}

## Initialization
作为网络安全漏洞分类专家，你必须遵守上述Rules，按照Workflows执行任务。

以下是需要分类的漏洞信息：
{chr(10).join(vuln_descriptions)}"""

        try:
            # 调用分类API
            print(f"    调用AI进行漏洞分类...")
            classification_result = analyze_code_assumptions(classification_prompt)
            classification_data = json.loads(self._extract_json_from_text(classification_result))
            
            print(f"    AI分类结果: {len(classification_data)} 个分组")
            for group_name, keys in classification_data.items():
                print(f"      {group_name}: {len(keys)} 个漏洞")
            
            # 根据分类结果构建新的分组
            new_groups = []
            for group_name, keys in classification_data.items():
                # 根据keys筛选对应的行
                group_rows = group[group['UUID'].astype(str).isin([str(k) for k in keys])]
                if not group_rows.empty:
                    new_groups.append(group_rows)
                    print(f"      创建分组 {group_name}: {len(group_rows)} 个漏洞")
            
            if not new_groups:
                print(f"    分类失败，返回原分组")
                return [group]
            
            print(f"    分类成功: {len(group)} -> {len(new_groups)} 个子组")
            return new_groups
            
        except Exception as e:
            print(f"    漏洞分类失败: {str(e)}")
            return [group]

    def _prepare_next_round_groups_simple(self, results_list):
        """简化版的下一轮分组准备（处理结果字典列表）"""
        print(f"    为下一轮重新分组 {len(results_list)} 个结果...")
        
        # 直接按顺序进行分组，不再排序
        print(f"    按顺序进行分组...")
        
        # 将结果列表按最大组大小进行分组
        next_round_groups = []
        
        # 使用多线程创建分组
        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
            group_futures = []
            
            for i in range(0, len(results_list), self.max_group_size):
                group_data = results_list[i:i + self.max_group_size]
                future = executor.submit(self._create_group, i, group_data, len(results_list))
                group_futures.append(future)
            
            with tqdm(total=len(group_futures), desc="创建下一轮分组") as pbar:
                for future in concurrent.futures.as_completed(group_futures):
                    try:
                        group_df = future.result()
                        next_round_groups.append(group_df)
                    except Exception as e:
                        print(f"创建分组失败: {str(e)}")
                    pbar.update(1)
        
        # 按索引排序确保顺序正确
        next_round_groups.sort(key=lambda x: x.index[0] if len(x) > 0 else 0)
        
        print(f"    为下一轮创建了 {len(next_round_groups)} 个分组")
        return next_round_groups

    def _create_group(self, start_index, group_data, total_len):
        """创建单个分组"""
        group_df = pd.DataFrame(group_data)
        
        start_idx = start_index + 1
        end_idx = min(start_index + self.max_group_size, total_len)
        
        print(f"      创建下一轮分组: 项目 {start_idx}-{end_idx} ({len(group_data)} 个)")
        
        return group_df

    def _build_final_results(self, final_groups):
        """构建最终结果（简化版，处理结果字典列表）"""
        print(f"\n=== 构建最终结果 ===")
        
        # 如果final_groups是字典列表，直接返回
        if isinstance(final_groups, list) and all(isinstance(item, dict) for item in final_groups):
            print(f"最终结果已是字典列表格式: {len(final_groups)} 个漏洞报告")
            print(f"  - 所有报告均为独立报告（已去重）")
            return final_groups
        
        # 否则按原逻辑处理
        print(f"处理 {len(final_groups)} 个最终分组...")
        all_results = []
        
        for i, group_result in enumerate(final_groups):
            if isinstance(group_result, dict):
                all_results.append(group_result)
                print(f"  最终分组 {i+1}: 字典类型 (1 个独立报告)")
            elif isinstance(group_result, pd.DataFrame):
                records = group_result.to_dict('records')
                all_results.extend(records)
                print(f"  最终分组 {i+1}: DataFrame类型 ({len(records)} 个独立报告)")
            else:
                if hasattr(group_result, 'to_dict'):
                    all_results.append(group_result.to_dict())
                    print(f"  最终分组 {i+1}: 其他类型转为字典 (1 个报告)")
        
        print(f"最终结果构建完成: {len(all_results)} 个漏洞报告")
        print(f"  - 所有报告均为独立报告（已去重）")
        
        return all_results

    def _clean_text_for_excel(self, text):
        """清理文本中的特殊字符，确保Excel兼容性"""
        if pd.isna(text):
            return ''
        
        # 移除或替换可能导致Excel问题的字符
        text = str(text).strip()
        # 替换常见的特殊字符
        replacements = {
            '\r': ' ',
            '\n': ' ',
            '\t': ' ',
            '\f': ' ',
            '\v': ' ',
            '\0': '',
            '\x00': '',
            '\u0000': '',
        }
        
        for old, new in replacements.items():
            text = text.replace(old, new)
        
        # 移除其他不可见字符
        text = ''.join(char for char in text if ord(char) >= 32 or char == ' ')
        
        return text

    def _extract_json_from_text(self, text):
        """从文本中提取JSON字符串，增加防御性编程措施"""
        print("\nDebug - Starting JSON extraction")
        print(f"Debug - Input text length: {len(text)}")
        
        try:
            # 首先尝试直接解析整个文本
            try:
                json.loads(text)
                return text
            except json.JSONDecodeError:
                pass
            
            # 策略1: 查找标准JSON代码块标记
            json_markers = ['```json', '`json', '```']
            for marker in json_markers:
                if marker in text:
                    # 找到标记后的内容
                    start_pos = text.find(marker) + len(marker)
                    end_marker = '```' if marker.startswith('```') else '`'
                    end_pos = text.find(end_marker, start_pos)
                    
                    if end_pos != -1:
                        json_candidate = text[start_pos:end_pos].strip()
                        try:
                            json.loads(json_candidate)
                            print(f"Debug - Found JSON in {marker} block")
                            return json_candidate
                        except json.JSONDecodeError:
                            continue
            
            # 策略2: 查找所有可能的JSON对象
            json_patterns = [
                r'\{"[^"]+"\s*:\s*\[[^\]]*\][^}]*\}',  # 标准格式 {"key":["val1","val2"]}
                r'\{[^{}]*"group_[^"]*"[^{}]*\}',      # 包含group_的对象
                r'\{[^{}]*:\s*\[[^\]]*\][^{}]*\}',     # 任何key:array格式
            ]
            
            for pattern in json_patterns:
                matches = re.findall(pattern, text, re.DOTALL)
                for match in matches:
                    try:
                        # 清理匹配结果
                        cleaned_match = match.strip()
                        json.loads(cleaned_match)
                        print(f"Debug - Found JSON with pattern: {pattern[:30]}...")
                        return cleaned_match
                    except json.JSONDecodeError:
                        continue
            
            # 策略3: 手动查找最后一个完整的JSON对象
            # 从文本末尾开始查找}，然后向前找到匹配的{
            last_brace = text.rfind('}')
            if last_brace != -1:
                # 从}位置向前查找匹配的{
                brace_count = 1
                start_pos = -1
                
                for i in range(last_brace - 1, -1, -1):
                    char = text[i]
                    if char == '}':
                        brace_count += 1
                    elif char == '{':
                        brace_count -= 1
                        if brace_count == 0:
                            start_pos = i
                            break
                
                if start_pos != -1:
                    json_candidate = text[start_pos:last_brace + 1]
                    try:
                        json.loads(json_candidate)
                        print(f"Debug - Found JSON by bracket matching")
                        return json_candidate
                    except json.JSONDecodeError:
                        pass
            
            # 策略4: 查找包含特定关键词的行，尝试提取
            lines = text.split('\n')
            for line in lines:
                line = line.strip()
                if 'group_' in line and '{' in line and '}' in line:
                    try:
                        json.loads(line)
                        print(f"Debug - Found JSON in line with group_")
                        return line
                    except json.JSONDecodeError:
                        continue
            
            # 策略4.5: 专门处理"步骤4"输出格式
            # 查找"步骤4"或类似标记后的JSON
            step_markers = ['步骤4', '**步骤4**', 'Step 4', '输出分类结果', '分类结果']
            for marker in step_markers:
                if marker in text:
                    # 找到标记位置
                    marker_pos = text.find(marker)
                    # 从标记后开始查找JSON
                    remaining_text = text[marker_pos:]
                    
                    # 在剩余文本中查找JSON
                    if '{' in remaining_text and '}' in remaining_text:
                        start = remaining_text.find('{')
                        end = remaining_text.rfind('}') + 1
                        json_candidate = remaining_text[start:end]
                        
                        try:
                            json.loads(json_candidate)
                            print(f"Debug - Found JSON after {marker}")
                            return json_candidate
                        except json.JSONDecodeError:
                            # 尝试提取最后一行的JSON
                            lines_after_marker = remaining_text[start:].split('\n')
                            for line in lines_after_marker:
                                line = line.strip()
                                if line.startswith('{') and line.endswith('}'):
                                    try:
                                        json.loads(line)
                                        print(f"Debug - Found JSON in line after {marker}")
                                        return line
                                    except json.JSONDecodeError:
                                        continue
            
            # 策略5: 尝试修复常见的JSON格式问题
            # 查找看起来像JSON的内容并尝试修复
            potential_json = None
            if '{' in text and '}' in text:
                start = text.find('{')
                end = text.rfind('}') + 1
                potential_json = text[start:end]
                
                # 尝试一些常见的修复
                fixes = [
                    lambda x: x,  # 原样
                    lambda x: x.replace("'", '"'),  # 单引号改双引号
                    lambda x: re.sub(r'(\w+):', r'"\1":', x),  # 给key加引号
                    lambda x: re.sub(r':\s*([^",\[\]{}]+)(?=[,}])', r': "\1"', x),  # 给value加引号
                ]
                
                for fix_func in fixes:
                    try:
                        fixed_json = fix_func(potential_json)
                        json.loads(fixed_json)
                        print(f"Debug - Fixed JSON format")
                        return fixed_json
                    except (json.JSONDecodeError, Exception):
                        continue
            
            raise ValueError("No valid JSON object found after all strategies")
            
        except (json.JSONDecodeError, ValueError) as e:
            print(f"Debug - Error extracting JSON: {str(e)}")
            print(f"Debug - Original text: {text}")
            
            # 最后的备用策略：尝试构造一个简单的分组
            print("Debug - Attempting fallback JSON construction")
            try:
                # 如果文本中提到了分组，尝试构造一个简单的JSON
                if 'group_' in text.lower():
                    fallback_json = '{"group_1":[]}'
                    print("Debug - Using fallback JSON")
                    return fallback_json
            except:
                pass
            
            raise ValueError(f"Failed to extract valid JSON after all strategies: {str(e)}")

    def _translate_to_chinese(self, final_results):
        """将漏洞结果翻译成中文"""
        if not self.enable_chinese_translation:
            print("中文翻译功能未启用，跳过翻译")
            return final_results
        
        print(f"\n=== 开始中文翻译 ===")
        print(f"需要翻译 {len(final_results)} 个漏洞结果...")
        
        translated_results = []
        
        with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
            future_to_index = {
                executor.submit(self._translate_single_result, i, result): i 
                for i, result in enumerate(final_results)
            }
            
            with tqdm(total=len(final_results), desc="中文翻译") as pbar:
                for future in concurrent.futures.as_completed(future_to_index):
                    try:
                        index = future_to_index[future]
                        translated_result = future.result()
                        
                        # 保持原有顺序
                        while len(translated_results) <= index:
                            translated_results.append(None)
                        translated_results[index] = translated_result
                        
                        print(f"  漏洞 {index+1} 翻译完成")
                    except Exception as e:
                        index = future_to_index[future]
                        original_result = final_results[index]
                        print(f"  漏洞 {index+1} 翻译失败: {str(e)}")
                        
                        # 保持原有顺序
                        while len(translated_results) <= index:
                            translated_results.append(None)
                        translated_results[index] = original_result
                    pbar.update(1)
        
        # 过滤掉None值并保持顺序
        translated_results = [result for result in translated_results if result is not None]
        
        print(f"中文翻译完成: {len(translated_results)} 个漏洞结果")
        return translated_results

    def _translate_single_result(self, index, result):
        """翻译单个漏洞结果"""
        original_result = result.get('漏洞结果', '')
        
        if not original_result or pd.isna(original_result):
            print(f"    漏洞 {index+1}: 无内容，跳过翻译")
            return result
        
        translate_prompt = f"""将这个漏洞详细的用中文解释一下，不要遗漏任何细节

漏洞描述：
{original_result}"""

        max_retries = 3  # 最大重试次数
        retry_count = 0
        
        while retry_count < max_retries:
            try:
                if retry_count == 0:
                    print(f"    漏洞 {index+1}: 开始翻译...")
                else:
                    print(f"    漏洞 {index+1}: 第 {retry_count+1} 次重试翻译...")
                
                translated_description = analyze_code_assumptions(translate_prompt)
                
                # 清理翻译结果
                cleaned_description = self._clean_text_for_excel(translated_description)
                
                # 检查翻译后的长度
                if len(cleaned_description) == 0:
                    print(f"    漏洞 {index+1}: ⚠️  翻译后长度为0，原长度 {len(original_result)}")
                    retry_count += 1
                    if retry_count < max_retries:
                        print(f"    漏洞 {index+1}: 准备重新翻译 (第 {retry_count+1} 次尝试)")
                        continue
                    else:
                        print(f"    漏洞 {index+1}: ❌ 重试 {max_retries} 次后仍然失败，保留原结果")
                        return result
                
                # 创建新的结果副本
                translated_result = result.copy()
                translated_result['漏洞结果'] = cleaned_description
                
                print(f"    漏洞 {index+1}: 翻译成功，原长度 {len(original_result)} -> 新长度 {len(cleaned_description)}")
                
                return translated_result
                
            except Exception as e:
                print(f"    漏洞 {index+1}: 翻译失败 - {str(e)}")
                retry_count += 1
                if retry_count < max_retries:
                    print(f"    漏洞 {index+1}: 准备重试 (第 {retry_count+1} 次尝试)")
                    continue
                else:
                    print(f"    漏洞 {index+1}: ❌ 重试 {max_retries} 次后仍然失败，保留原结果")
                    return result
        
        # 如果所有重试都失败，返回原结果
        return result

    @staticmethod
    def perform_post_reasoning_deduplication(project_id, db_engine, logger):
        """在reasoning完成后，validation开始前进行去重处理"""
        from logging_config import log_step, log_section_start, log_section_end, log_error, log_warning, log_success, log_data_info
        from dao import ProjectTaskMgr
        
        log_step(logger, "开始获取reasoning后的漏洞数据")
        
        try:
            # 获取reasoning后的所有漏洞数据
            project_taskmgr = ProjectTaskMgr(project_id, db_engine)
            entities = project_taskmgr.query_task_by_project_id(project_id)
            
            # 调试信息：统计所有实体
            total_entities = len(entities)
            log_data_info(logger, "总任务实体数量", total_entities)
            print(f"🔍 调试信息 - 总任务实体数量: {total_entities}")
            
            # 筛选有漏洞结果的数据
            vulnerability_data = []
            for entity in entities:
                # 调试每个实体的详细信息
                has_result = bool(entity.result)
                has_business_code = hasattr(entity, 'business_flow_code') and entity.business_flow_code and len(entity.business_flow_code) > 0
                
                if has_result and has_business_code:
                    vulnerability_data.append({
                        '漏洞结果': entity.result,
                        'ID': entity.id,
                        '项目名称': entity.project_id,
                        '合同编号': entity.contract_code,
                        'UUID': entity.uuid,
                        '函数名称': entity.name,
                        '函数代码': entity.content,
                        '规则类型': entity.rule_key,
                        '开始行': entity.start_line,
                        '结束行': entity.end_line,
                        '相对路径': entity.relative_file_path,
                        '绝对路径': entity.absolute_file_path,
                        '业务流程代码': entity.business_flow_code,
                        '扫描记录': entity.scan_record,
                        '推荐': entity.recommendation
                    })
            
            filtered_count = len(vulnerability_data)
            print(f"🔍 调试信息 - 通过筛选条件的实体: {filtered_count}")
                        
            original_df = pd.DataFrame(vulnerability_data)
            original_count = len(original_df)
            original_ids = set(original_df['ID'].astype(str))
            
            log_data_info(logger, "去重前漏洞数量", original_count)
            log_data_info(logger, "去重前漏洞ID", f"{', '.join(sorted(original_ids))}")
            
            # 检查是否存在已经被逻辑删除的记录（short_result='delete'）
            all_entities = project_taskmgr.query_task_by_project_id(project_id)
            deleted_tasks = [entity for entity in all_entities if getattr(entity, 'short_result', '') == 'delete']
            
            if deleted_tasks:
                deleted_count = len(deleted_tasks)
                deleted_ids = [str(task.id) for task in deleted_tasks]
                print(f"\n⚠️  检测到已有 {deleted_count} 个逻辑删除的记录，跳过ResProcessor去重处理")
                print(f"已删除的ID: {', '.join(deleted_ids)}")
                log_warning(logger, f"跳过ResProcessor去重处理 - 检测到{deleted_count}个已逻辑删除的记录")
                log_warning(logger, f"已删除的ID: {', '.join(deleted_ids)}")
                return
            
            # 使用ResProcessor进行去重
            log_step(logger, "开始ResProcessor去重处理")
            res_processor = ResProcessor(original_df, max_group_size=5, iteration_rounds=8, enable_chinese_translation=False)
            processed_df = res_processor.process()
            
            deduplicated_count = len(processed_df)
            deduplicated_ids = set(processed_df['ID'].astype(str))
            
            log_data_info(logger, "去重后漏洞数量", deduplicated_count)
            log_data_info(logger, "去重后漏洞ID", f"{', '.join(sorted(deduplicated_ids))}")
            
            # 计算被去重的ID
            removed_ids = original_ids - deduplicated_ids
            removed_count = len(removed_ids)
            
            # 打印去重结果
            print(f"\n{'='*60}")
            print(f"🔄 Reasoning后去重处理结果")
            print(f"{'='*60}")
            print(f"去重前漏洞数量: {original_count}")
            print(f"去重后漏洞数量: {deduplicated_count}")
            print(f"被去重的漏洞数量: {removed_count}")
            
            if removed_ids:
                print(f"\n🗑️  被去重的漏洞ID列表:")
                for i, removed_id in enumerate(sorted(removed_ids), 1):
                    print(f"  {i:2d}. ID: {removed_id}")
                
                # 逻辑删除被去重的记录 - 将short_result设置为"delete"
                print(f"\n🗑️  开始逻辑删除被去重的记录(设置short_result='delete')...")
                marked_count = 0
                failed_marks = []
                
                for removed_id in removed_ids:
                    try:
                        # 转换为整数类型的ID
                        id_int = int(removed_id)
                        project_taskmgr.update_short_result(id_int, "delete")
                        marked_count += 1
                        print(f"    ✅ 标记成功: ID {removed_id} -> short_result='delete'")
                    except Exception as e:
                        failed_marks.append(removed_id)
                        print(f"    ❌ 标记出错: ID {removed_id}, 错误: {str(e)}")
                        logger.error(f"标记删除ID {removed_id} 时出错: {str(e)}")
                
                print(f"\n📊 逻辑删除结果:")
                print(f"    成功标记: {marked_count} 条记录")
                if failed_marks:
                    print(f"    标记失败: {len(failed_marks)} 条记录 - IDs: {', '.join(failed_marks)}")
                    logger.warning(f"标记失败的IDs: {', '.join(failed_marks)}")
                
                log_success(logger, "逻辑删除操作完成", f"成功标记: {marked_count}/{removed_count}")
            else:
                print("✅ 没有漏洞被去重")
            
            print(f"{'='*60}\n")
            
            # 记录到日志
            log_success(logger, "去重处理完成", f"原始: {original_count} -> 去重后: {deduplicated_count}, 逻辑删除: {removed_count}")
            if removed_ids:
                logger.info(f"被去重的漏洞ID: {', '.join(sorted(removed_ids))}")
                logger.info(f"逻辑删除了 {marked_count} 条被去重的记录(设置short_result='delete')")
            
        except Exception as e:
            log_error(logger, "去重处理失败", e)
            import traceback
            logger.error(f"详细错误信息: {traceback.format_exc()}")

    @staticmethod
    def generate_excel(output_path, project_id, db_engine):
        """生成Excel报告"""
        from dao import ProjectTaskMgr
        
        project_taskmgr = ProjectTaskMgr(project_id, db_engine)
        entities = project_taskmgr.query_task_by_project_id(project_id)
        
        # 创建一个空的DataFrame来存储所有实体的数据
        data = []
        total_entities = len(entities)
        deleted_entities = 0
        
        for entity in entities:
            # 跳过已逻辑删除的记录
            if getattr(entity, 'short_result', '') == 'delete':
                deleted_entities += 1
                continue
                
            # 优先使用validation后的short_result，如果没有则使用原始result
            short_result = entity.short_result
            result = entity.result
            if short_result and ("yes" in str(short_result).lower()) and len(entity.business_flow_code)>0:
                # Extract funds_at_risk from scan_record if available
                far_value = None
                try:
                    scan_data = json.loads(entity.scan_record) if entity.scan_record else {}
                    far_value = scan_data.get('funds_at_risk', None)
                except Exception:
                    far_value = None

                data.append({
                    '漏洞结果': result,
                    'ID': entity.id,
                    '项目名称': entity.project_id,
                    '合同编号': entity.contract_code,
                    'UUID': entity.uuid,  # 使用uuid而不是key
                    '函数名称': entity.name,
                    '函数代码': entity.content,
                    '规则类型': entity.rule_key,  # 新增rule_key
                    '开始行': entity.start_line,
                    '结束行': entity.end_line,
                    '相对路径': entity.relative_file_path,
                    '绝对路径': entity.absolute_file_path,
                    '业务流程代码': entity.business_flow_code,
                    '扫描记录': entity.scan_record,  # 使用新的scan_record字段
                    '推荐': entity.recommendation,
                    'funds_at_risk': far_value if far_value is not None else 0
                })
        
        # 打印数据统计信息
        print(f"\n📊 Excel报告数据统计:")
        print(f"   总记录数: {total_entities}")
        print(f"   逻辑删除的记录数: {deleted_entities}")
        print(f"   有效记录数: {total_entities - deleted_entities}")
        print(f"   符合条件的漏洞记录数: {len(data)}")
        
        # 将数据转换为DataFrame
        if not data:  # 检查是否有数据
            print("No data to process")
            return
            
        df = pd.DataFrame(data)
        
        try:
            # 对df进行漏洞归集处理
            res_processor = ResProcessor(df, max_group_size=5, iteration_rounds=5, enable_chinese_translation=True)
            processed_df = res_processor.process()
            
            # 确保所有必需的列都存在
            required_columns = df.columns
            for col in required_columns:
                if col not in processed_df.columns:
                    processed_df[col] = ''
                    
            # 重新排列列顺序以匹配原始DataFrame
            processed_df = processed_df[df.columns]
        except Exception as e:
            print(f"Error processing data: {e}")
            processed_df = df  # 如果处理失败，使用原始DataFrame
        
        # 确保输出目录存在
        output_dir = os.path.dirname(output_path)
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        # 检查文件是否存在，如果不存在则创建新文件
        if not os.path.exists(output_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "项目数据"
        else:
            wb = load_workbook(output_path)
            if "项目数据" in wb.sheetnames:
                ws = wb["项目数据"]
            else:
                ws = wb.create_sheet("项目数据")
        
        # 如果工作表是空的，添加表头
        if ws.max_row == 1:
            for col, header in enumerate(processed_df.columns, start=1):
                ws.cell(row=1, column=col, value=header)
        
        # 将DataFrame数据写入工作表
        for row in dataframe_to_rows(processed_df, index=False, header=False):
            ws.append(row)
        
        # 保存Excel文件
        wb.save(output_path)
        
        print(f"Excel文件已保存到: {output_path}")